from __future__ import annotations

import logging
import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.drawing.image import Image as XLImage
from pandas.api.types import (
    is_datetime64_any_dtype,
    is_datetime64tz_dtype,
    is_object_dtype,
)
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas

from .models import AnalysisArtifact
from .validation import quality_report_to_frames


LOGGER = logging.getLogger(__name__)
INVALID_SHEET_TITLE_CHARS = re.compile(r"[:\\/?*\[\]]")
EXCEL_CELL_CHAR_LIMIT = 32767


def _excel_safe_sheet_name(name: str) -> str:
    cleaned = INVALID_SHEET_TITLE_CHARS.sub("_", str(name))
    cleaned = cleaned.strip("'")
    return cleaned[:31] or "sheet"


def _sanitize_for_excel(frame: pd.DataFrame) -> pd.DataFrame:
    def clean_value(value):
        if isinstance(value, bytes):
            value = value.decode("utf-8", errors="replace")
        if isinstance(value, pd.Timestamp) and value.tzinfo is not None:
            value = value.tz_localize(None)
        if isinstance(value, str):
            value = ILLEGAL_CHARACTERS_RE.sub("", value)
            return value[:EXCEL_CELL_CHAR_LIMIT]
        return value

    sanitized = frame.copy()
    sanitized.columns = [clean_value(column) for column in sanitized.columns]
    for column in sanitized.columns:
        if is_datetime64tz_dtype(sanitized[column]):
            sanitized[column] = sanitized[column].dt.tz_localize(None)
            continue
        if is_datetime64_any_dtype(sanitized[column]) and not is_object_dtype(sanitized[column]):
            continue
        sanitized[column] = sanitized[column].map(clean_value)
    return sanitized


def export_excel_report(artifact: AnalysisArtifact, output_dir: Path) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    excel_path = output_dir / "threat_intelligence_report.xlsx"
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        _sanitize_for_excel(artifact.cleaned).to_excel(writer, sheet_name="cleaned_dataset", index=False)
        for name, frame in quality_report_to_frames(artifact.quality_report).items():
            _sanitize_for_excel(frame).to_excel(writer, sheet_name=_excel_safe_sheet_name(name), index=False)
        pd.DataFrame(
            [
                {
                    "row_count": artifact.summary_statistics["row_count"],
                    "column_count": artifact.summary_statistics["column_count"],
                }
            ]
        ).pipe(_sanitize_for_excel).to_excel(writer, sheet_name="summary", index=False)
        pd.DataFrame({"executive_summary": [artifact.executive_summary]}).pipe(_sanitize_for_excel).to_excel(
            writer, sheet_name="insights", index=False
        )
        pd.DataFrame({"recommendation": artifact.recommendations}).pipe(_sanitize_for_excel).to_excel(
            writer, sheet_name="recommendations", index=False
        )
        pd.DataFrame({"transformation": artifact.transformation_log}).pipe(_sanitize_for_excel).to_excel(
            writer, sheet_name="transformations", index=False
        )
        for sheet_name, frame in artifact.segmentations.items():
            _sanitize_for_excel(frame).to_excel(
                writer,
                sheet_name=_excel_safe_sheet_name(f"seg_{sheet_name}"),
                index=False,
            )
        for sheet_name, frame in artifact.trends.items():
            _sanitize_for_excel(frame).to_excel(
                writer,
                sheet_name=_excel_safe_sheet_name(f"trend_{sheet_name}"),
                index=False,
            )
        if not artifact.correlations.empty:
            _sanitize_for_excel(artifact.correlations).to_excel(
                writer, sheet_name=_excel_safe_sheet_name("correlations")
            )
        if not artifact.anomalies.empty:
            _sanitize_for_excel(artifact.anomalies).to_excel(
                writer,
                sheet_name=_excel_safe_sheet_name("anomalies"),
                index=False,
            )

    workbook = load_workbook(excel_path)
    chart_sheet = workbook.create_sheet("charts")
    anchor_row = 1
    for name, path in artifact.generated_files.items():
        if path.suffix.lower() != ".png":
            continue
        chart_sheet[f"A{anchor_row}"] = name
        image = XLImage(str(path))
        image.anchor = f"A{anchor_row + 1}"
        chart_sheet.add_image(image)
        anchor_row += 22
    workbook.save(excel_path)
    LOGGER.info("Exported Excel report to %s", excel_path)
    return excel_path


def export_pdf_report(artifact: AnalysisArtifact, output_dir: Path) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    pdf_path = output_dir / "threat_intelligence_summary.pdf"
    report = canvas.Canvas(str(pdf_path), pagesize=letter)
    width, height = letter
    y = height - 50

    def write_line(text: str, font: str = "Helvetica", size: int = 11) -> None:
        nonlocal y
        if y < 50:
            report.showPage()
            y = height - 50
        report.setFont(font, size)
        report.drawString(50, y, text[:110])
        y -= 16

    write_line("Threat Intelligence Analytics Summary", font="Helvetica-Bold", size=16)
    write_line("")
    write_line("Executive Summary", font="Helvetica-Bold", size=13)
    for sentence in artifact.executive_summary.split(". "):
        if sentence.strip():
            write_line(f"- {sentence.strip()}")
    write_line("")
    write_line("Recommended Analyst Actions", font="Helvetica-Bold", size=13)
    for item in artifact.recommendations:
        write_line(f"- {item}")
    write_line("")
    write_line("Data Quality Snapshot", font="Helvetica-Bold", size=13)
    write_line(f"Rows: {artifact.quality_report.total_rows}")
    write_line(f"Columns: {artifact.quality_report.total_columns}")
    write_line(f"Duplicate Rows: {artifact.quality_report.duplicate_rows}")
    write_line(f"Malformed Rows: {artifact.quality_report.malformed_rows}")
    write_line("")
    write_line("Generated Charts", font="Helvetica-Bold", size=13)
    for name, path in artifact.generated_files.items():
        if path.suffix.lower() == ".png":
            write_line(f"- {name}: {path.name}")
    report.save()
    LOGGER.info("Exported PDF report to %s", pdf_path)
    return pdf_path
